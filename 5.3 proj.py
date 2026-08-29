from datetime import datetime
import csv
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference

print("AMAHOP2918's Spreadsheet Automation Menu")

# List for the options
menuOptions = [
    "1 Input Data",
    "2 View Current Data",
    "3 Generate Report"
]

# Getting user menu option
print("Choose a number from the following options")
for option in menuOptions:
    print(option)

choice = input("Enter your choice: ")

# F to C conversion
def convertData(data):
    return (data - 32) * 5 / 9

# Appends comma-separated data to a CSV file or creates the file if needed.
def insertData(filePath, data):
    try:
        with open(filePath, "a") as file:
            file.write(data + "\n")
        return True
    except Exception as error:
        print("Error writing to file:", error)
        return False

# Displays the contents of the CSV file and the path being read.
def viewData(filePath):
    try:
        print("Reading data from:", filePath)

        with open(filePath, "r") as file:
            contents = file.read()
            print(contents)

    except Exception as error:
        print("Error reading file:", error)

# Getting information that will be used from the menu selection and putting it in the
# csv file.
def getInput():
    try:
        entries = int(input("How many entries are you inputting?\n"))

        for i in range(entries):
            entryDate = input("Enter a date:\n")
            temperature = float(
                input("Enter the highest temp for the inputted date:\n")
            )

            # Value that takes the inputted temp for conversion.
            convertedValue = convertData(temperature)

            data = entryDate + "," + str(temperature) + "," + str(convertedValue)

            if insertData("ZooData.csv", data):
                print(
                    "The following data was saved at "
                    + str(datetime.now())
                    + ": "
                    + data
                    + "."
                )

    except Exception as error:
        print("Error entering data:", error)

# createChart requires a CSV file path (str) and chart type (str: "line" or "bar").
# It asks the user to select Fahrenheit or Celsius data, saves the selected data
# to final.xlsx, and creates a chart using the dates as labels.
# Returns None.
def createChart(filePath, chartType):
    try:
        # Ask the user which data source they want to use.
        print("\nChoose the data source for the chart:")
        print("1 Fahrenheit (initial data)")
        print("2 Celsius (converted data)")

        dataChoice = input("Enter your choice: ")

        # Determine which CSV column and Excel column to use.
        if dataChoice == "1":
            dataColumn = 2
            dataName = "Fahrenheit"
            yAxisLabel = "Temperature (°F)"
        elif dataChoice == "2":
            dataColumn = 3
            dataName = "Celsius"
            yAxisLabel = "Temperature (°C)"
        else:
            print("Error: Invalid data source selected.")
            return

        # Open and read the CSV file.
        dates = []
        values = []

        with open(filePath, "r") as file:
            reader = csv.reader(file)

            for row in reader:
                if len(row) >= 3:
                    dates.append(row[0])
                    values.append(float(row[dataColumn - 1]))

        # Create a new Excel workbook.
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Temperature Data"

        # Add headers.
        worksheet["A1"] = "Date"
        worksheet["B1"] = dataName

        # Add dates and selected temperature data.
        for i in range(len(dates)):
            worksheet.cell(row=i + 2, column=1, value=dates[i])
            worksheet.cell(row=i + 2, column=2, value=values[i])

        # Create the appropriate chart.
        if chartType == "line":
            chart = LineChart()
        elif chartType == "bar":
            chart = BarChart()
        else:
            print("Error: Invalid chart type.")
            return

        # Select the temperature values for the chart.
        data = Reference(
            worksheet,
            min_col=2,
            min_row=1,
            max_row=len(values) + 1
        )

        # Select the dates for the x-axis labels.
        categories = Reference(
            worksheet,
            min_col=1,
            min_row=2,
            max_row=len(dates) + 1
        )

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Chart title and axis labels.
        currentDate = datetime.now().strftime("%m/%d/%Y")
        chart.title = "AMAHOP2918 " + currentDate
        chart.x_axis.title = "Date"
        chart.y_axis.title = yAxisLabel

        # Add the chart to the worksheet.
        worksheet.add_chart(chart, "D2")

        # Save the Excel spreadsheet.
        workbook.save("final.xlsx")

        print("\nReport successfully created!")
        print("The selected data was saved to final.xlsx.")

    except Exception as error:
        print("Error creating report:", error)


# generateReport requires a CSV file path (str).
# It asks the user to choose a line or bar chart and calls createChart
# using the selected chart type.
# Returns None.
def generateReport(filePath):
    print("\nChoose the graph type you would like to create:")
    print("1 Line chart")
    print("2 Bar chart")

    graphChoice = input("Enter your choice: ")

    if graphChoice == "1":
        createChart(filePath, "line")
    elif graphChoice == "2":
        createChart(filePath, "bar")
    else:
        print("Error: Invalid graph type selected.")

# Displays if choice is valid and date.
if choice in ["1", "2", "3"]:
    print("You selected", choice, "at", datetime.now())
else:
    print("Error: Invalid choice selected.")

# Runs Input Data functionality if option 1 was selected.
if choice == "1":
    getInput()

# Runs View Current Data functionality if option 2 was selected.
elif choice == "2":
    viewData("ZooData.csv")

# Runs Generate Report functionality if option 3 was selected.
elif choice == "3":
    generateReport("ZooData.csv")
